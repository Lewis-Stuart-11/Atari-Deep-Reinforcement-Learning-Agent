# Author Lewis Stuart 201262348

import matplotlib.pyplot as plt
import numpy as np
import os

from optimal_game_parameters import optimal_game_parameters

# Writes out the essential information of the training episodes to an excel file for each game
def write_final_results(info_per_episode, default_atari_game, num_training_episodes, agent_parameters):
    import pandas as pd
    import xlsxwriter
    from openpyxl import load_workbook

    # Dataframe of the results of each episode
    results_data_frame = pd.DataFrame(info_per_episode)
    results_data_frame.name = "AI results"

    # Converts parameter data into a dataframe
    parameter_data_dict = agent_parameters._asdict()
    for key in parameter_data_dict.keys():
        parameter_data_dict[key] = str(parameter_data_dict[key])

    parameter_data_dict["num_training_episodes"] = str(num_training_episodes)

    parameter_data_frame = pd.DataFrame(parameter_data_dict, index=[0])
    parameter_data_frame.name = "Parameter information"

    # Sets Excel file as the atari game
    file_name = "results/" + default_atari_game + "_results.xlsx"

    # If the file does not exist currently, then it is created
    if not os.path.exists(file_name):
        workbook = xlsxwriter.Workbook(file_name)
        workbook.close()

    # File is opened and dataframes are added to a new sheet
    workbook = load_workbook(file_name)
    with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
        writer.book = workbook
        parameter_data_frame.to_excel(writer, sheet_name='Results_', startrow=1, startcol=0)
        results_data_frame.to_excel(writer, sheet_name='Results_', startrow=parameter_data_frame.shape[0] + 5, startcol=0)
        writer.save()


# Plots the current episode statistics
def plot(info_per_episode, save_plot):

    # Declares the number of steps, total reward and total time for each episode
    steps_per_episode = [episode["num_steps"] for episode in info_per_episode]
    rewards_per_episode = [round(episode["total_reward"], 2) for episode in info_per_episode]
    total_time = [float(episode["total_time"]) for episode in info_per_episode]
    moving_average = [episode["moving_average"] for episode in info_per_episode]
    epsilon = [float(episode["epsilon"])*100 for episode in info_per_episode] \
        if "epsilon" in info_per_episode[0].keys() else None

    num_episodes = len(steps_per_episode)

    # Sets up Rewards graph
    plt.figure(4)
    plt.clf()

    # Plot reward
    plt.plot(rewards_per_episode, '-rx', label="Rewards")

    # Plots moving averages
    plt.plot(moving_average, '-kx', label=f"Moving average")

    plt.title("Reward for each training episode")
    plt.xlabel("Episode number")
    plt.ylabel("Reward value")
    plt.legend(["Reward", "Moving average (reward)"])

    plt.locator_params(axis='y', nbins=5)
    plt.locator_params(axis='x', nbins=5)
    point_intervals = round(len(info_per_episode) / 10)
    if point_intervals < 1:
        point_intervals = 1
    plt.xticks(np.arange(0, len(info_per_episode), point_intervals))

    # Saves the final reward plot
    if save_plot:
        plt.savefig(f"results/Reward_{num_episodes}.png", dpi=300, bbox_inches='tight')
        #plt.show()

    plt.close()

    plt.pause(0.001)

    # Sets up main graph
    plt.figure(2)
    plt.clf()

    # Plots properties
    plt.plot(steps_per_episode, '-bx', label="Steps")
    plt.plot(rewards_per_episode, '-rx', label="Rewards")
    plt.plot(total_time, '-gx', label="Time")
    if epsilon:
        plt.plot(epsilon, '-yx', label="Epsilon")

    # Descriptors
    plt.title("Properties for each training episode")
    plt.xlabel("Episode number")
    plt.ylabel("Total per element")
    if epsilon:
        plt.legend(["Steps", "Reward", "Time (ms)", "Epsilon"])
    else:
        plt.legend(["Steps", "Reward", "Time (ms)"])

    plt.locator_params(axis='y', nbins=5)
    plt.locator_params(axis='x', nbins=5)

    point_intervals = round(len(info_per_episode) / 10)
    if point_intervals < 1:
        point_intervals = 1
    plt.xticks(np.arange(0, len(info_per_episode), point_intervals))

    # Saves the final main plot
    if save_plot:
        plt.savefig(f"results/Properties_{num_episodes}.png", dpi=300, bbox_inches='tight')
        #plt.show()

    plt.close()

    plt.pause(0.001)

